"""Streamlit Web UI for 価格演算システム.

起動: streamlit run src/price/app.py
"""
from __future__ import annotations

import io

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from price.config import load_config
from price.db.pool import PoolManager
from price.export.excel_writer import COLUMNS, write_results
from price.main import process_parts
from price.models.enums import PartPrefix, classify_prefix

# ---------- ページ設定 ----------
st.set_page_config(page_title="価格演算", layout="wide")

# ---------- カスタムCSS ----------
st.markdown("""
<style>
/* ヘッダー */
.main-header {
    background: linear-gradient(135deg, #1a1a2e 0%, #16213e 50%, #0f3460 100%);
    padding: 1.5rem 2rem;
    border-radius: 12px;
    margin-bottom: 1.5rem;
    box-shadow: 0 4px 15px rgba(0,0,0,0.2);
}
.main-header h1 {
    color: #e0e0e0;
    font-size: 1.8rem;
    font-weight: 700;
    margin: 0;
    letter-spacing: 0.05em;
}
/* メトリクスカード */
[data-testid="stMetric"] {
    background: linear-gradient(135deg, #f8f9fa 0%, #e9ecef 100%);
    border: 1px solid #dee2e6;
    border-radius: 10px;
    padding: 1rem;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06);
}
[data-testid="stMetric"] label {
    color: #495057;
    font-weight: 600;
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 0.05em;
}
[data-testid="stMetric"] [data-testid="stMetricValue"] {
    color: #1a1a2e;
    font-weight: 700;
}

/* セクションヘッダー */
.section-header {
    background: linear-gradient(90deg, #0f3460, #16213e);
    color: #e0e0e0;
    padding: 0.6rem 1.2rem;
    border-radius: 8px;
    font-size: 1rem;
    font-weight: 600;
    margin: 1rem 0 0.8rem 0;
    letter-spacing: 0.03em;
}

/* ボタンスタイル */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #0f3460 0%, #1a1a2e 100%);
    border: none;
    border-radius: 8px;
    font-weight: 600;
    letter-spacing: 0.05em;
    transition: all 0.3s ease;
    box-shadow: 0 2px 10px rgba(15,52,96,0.3);
}
.stButton > button[kind="primary"]:hover {
    box-shadow: 0 4px 20px rgba(15,52,96,0.5);
    transform: translateY(-1px);
}

/* タブ */
.stTabs [data-baseweb="tab-list"] {
    gap: 0;
}
.stTabs [data-baseweb="tab"] {
    border-radius: 8px 8px 0 0;
    padding: 0.6rem 1.5rem;
    font-weight: 600;
}

/* データフレーム */
[data-testid="stDataFrame"] {
    border-radius: 10px;
    overflow: hidden;
    box-shadow: 0 2px 10px rgba(0,0,0,0.08);
}

/* サイドバー */
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%);
}
section[data-testid="stSidebar"] .stMarkdown {
    color: #c0c0d0;
}

/* 仕切り線 */
hr {
    border: none;
    height: 2px;
    background: linear-gradient(90deg, transparent, #0f3460, transparent);
    margin: 1.5rem 0;
}

/* expander */
.streamlit-expanderHeader {
    font-weight: 600;
    color: #1a1a2e;
}
</style>
""", unsafe_allow_html=True)

# ---------- サイドバー: 設定（パスワード保護） ----------
_SETTINGS_PASS = "0018"

# デフォルト値（パスワード未入力時に使用）
settings_path = "config/settings.yaml"
rates_path = "config/rates.yaml"
rates_excel_path = None
rates_sheet_name = "テーブル"
a_assembly_mode = "simple"

with st.sidebar:
    st.markdown("### 設定")
    pw = st.text_input("設定パスワード", type="password", key="sidebar_pw")
    if pw == _SETTINGS_PASS:
        st.success("認証OK")
        settings_path = st.text_input("DB設定ファイル", value="config/settings.yaml")

        rate_source = st.radio("掛率ソース", ["YAML", "Excel"])
        if rate_source == "YAML":
            rates_path = st.text_input("掛率YAML", value="config/rates.yaml")
            rates_excel_path = None
            rates_sheet_name = "テーブル"
        else:
            rates_excel_path = st.text_input("掛率Excelファイル", value="config/掛率.xlsx")
            rates_sheet_name = st.text_input("シート名", value="テーブル")
            rates_path = None

        st.markdown("---")
        a_mode_label = st.selectbox(
            "A番組立金額",
            ["簡易式", "工数反映式", "工数反映式_2026新型"],
            key="a_assembly_mode_select",
        )
        _mode_map = {"簡易式": "simple", "工数反映式": "kousuu", "工数反映式_2026新型": "kousuu_2026"}
        a_assembly_mode = _mode_map[a_mode_label]
    elif pw:
        st.error("パスワードが違います")

# ---------- ヘッダー ----------
st.markdown("""
<div class="main-header">
    <h1>価格演算システム</h1>
</div>
""", unsafe_allow_html=True)


# ---------- ユーティリティ ----------
def _is_a_part(buhin_bango: str) -> bool:
    """A番かどうか判定する."""
    try:
        return classify_prefix(buhin_bango) == PartPrefix.A
    except ValueError:
        return False


def _is_m_part(buhin_bango: str) -> bool:
    """M番かどうか判定する."""
    try:
        return classify_prefix(buhin_bango) == PartPrefix.M
    except ValueError:
        return False


# ---------- 一括クリアコールバック ----------
def _clear_inputs():
    """手入力フィールドとアップロードファイルをクリアする."""
    st.session_state["manual_input_area"] = ""
    # file_uploader は session_state から直接クリアできないため、
    # キーサフィックスをインクリメントしてウィジェットを再生成する
    st.session_state["uploader_key"] = st.session_state.get("uploader_key", 0) + 1
    if "results" in st.session_state:
        del st.session_state["results"]
    if "stats" in st.session_state:
        del st.session_state["stats"]


# ---------- 入力タブ ----------
_uploader_key = st.session_state.get("uploader_key", 0)

tab_manual, tab_excel = st.tabs(["手入力", "Excelアップロード"])

with tab_excel:
    uploaded_file = st.file_uploader(
        "品番リスト（A列に品番、1行目ヘッダー）",
        type=["xlsx"],
        key=f"excel_uploader_{_uploader_key}",
    )

with tab_manual:
    manual_input = st.text_area(
        "品番を1行ずつ入力してください",
        height=200,
        placeholder="M12345\n4ABCDE\nA99999",
        key="manual_input_area",
    )


def _get_part_numbers() -> list[str] | None:
    """アクティブな入力方法から品番リストを取得する."""
    if uploaded_file is not None:
        wb = load_workbook(uploaded_file, read_only=True, data_only=True)
        ws = wb.active
        parts = []
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            val = row[0]
            if val is not None and str(val).strip():
                parts.append(str(val).strip())
        wb.close()
        return parts if parts else None
    if manual_input and manual_input.strip():
        lines = manual_input.strip().splitlines()
        return [ln.strip() for ln in lines if ln.strip()]
    return None


# ---------- ボタン ----------
btn_col1, btn_col2 = st.columns([3, 1])
with btn_col1:
    run_clicked = st.button("▶ 実行", type="primary", use_container_width=True)
with btn_col2:
    st.button("クリア", on_click=_clear_inputs, use_container_width=True)

if run_clicked:
    part_numbers = _get_part_numbers()
    if part_numbers is None:
        st.error("品番が入力されていません。Excelをアップロードするか、手入力してください。")
        st.stop()

    st.info(f"品番数: {len(part_numbers)}")

    # 設定読み込み
    try:
        config = load_config(
            settings_path,
            rates_path,
            rates_excel_path=rates_excel_path,
            rates_sheet_name=rates_sheet_name,
            a_assembly_mode=a_assembly_mode,
        )
    except Exception as e:
        st.error(f"設定読み込みエラー: {e}")
        st.stop()

    # DB接続プール初期化（冪等）
    try:
        PoolManager.init(config.eco_db, config.honps_db)
    except Exception as e:
        st.error(f"DB接続エラー: {e}")
        st.stop()

    # 処理実行（プログレスバー付き）
    progress_bar = st.progress(0, text="処理開始...")

    def _on_progress(step: int, total: int, msg: str) -> None:
        progress_bar.progress(step / total, text=msg)

    try:
        results, stats = process_parts(part_numbers, config, _on_progress)
    except Exception as e:
        st.error(f"処理エラー: {e}")
        st.stop()

    progress_bar.progress(1.0, text="完了")

    # セッションに保存（Streamlit の rerun でも保持）
    st.session_state["results"] = results
    st.session_state["stats"] = stats

# ---------- 結果表示 ----------
if "results" in st.session_state:
    results = st.session_state["results"]
    stats = st.session_state["stats"]

    # サマリー指標
    st.markdown('<div class="section-header">処理結果サマリー</div>',
                unsafe_allow_html=True)
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("処理件数", stats["total"])
    col2.metric("NULLデータ", stats["null_count"])
    col3.metric("データ取得", f"{stats['fetch_time']:.1f}秒")
    col4.metric("計算処理", f"{stats['calc_time']:.1f}秒")

    # デバッグ情報
    all_debug = stats.get("hyotanka_debug", []) + stats.get("kakaku_debug", [])
    if all_debug:
        with st.expander("データ取得ログ (デバッグ)"):
            for line in all_debug:
                st.text(line)

    # DataFrame 構築
    rows = []
    for r in results:
        row = {}
        for header, attr in COLUMNS:
            val = getattr(r, attr, None)
            if attr == "has_null_data":
                val = "有" if val else ""
            if attr == "standard_price" and val is not None:
                val = int(val)
            row[header] = val
        rows.append(row)
    df = pd.DataFrame(rows)

    # A番でNULLデータありの品番セットを作成
    a_null_parts = {r.buhin_bango for r in results
                    if _is_a_part(r.buhin_bango) and r.has_null_data}

    # 条件付き色付け
    def _highlight_rows(row_series: pd.Series) -> list[str]:
        styles = [""] * len(row_series)
        idx = {col: i for i, col in enumerate(row_series.index)}

        buhin = row_series.get("品番")

        # A番でNULLデータあり → 行全体を赤文字
        if buhin in a_null_parts:
            styles = ["color: #FF0000"] * len(row_series)

        std_price = row_series.get("標準単価")
        if std_price is None or std_price == 0:
            bg_yellow = "background-color: #FFFF00; color: #000000"
            styles[idx["品番"]] = styles[idx["品番"]] + "; " + bg_yellow if styles[idx["品番"]] else bg_yellow
            styles[idx["標準単価"]] = styles[idx["標準単価"]] + "; " + bg_yellow if styles[idx["標準単価"]] else bg_yellow

        flag = row_series.get("価格比較")
        if flag == "高":
            hi_style = "background-color: #FFFF00; color: #000000"
            styles[idx["H仕切り"]] = styles[idx["H仕切り"]] + "; " + hi_style if styles[idx["H仕切り"]] else hi_style
        elif flag == "安":
            lo_style = "background-color: #00FFFF; color: #000000"
            styles[idx["H仕切り"]] = styles[idx["H仕切り"]] + "; " + lo_style if styles[idx["H仕切り"]] else lo_style

        # 部品区分が2または4を含む場合は赤塗り
        kubun = row_series.get("部品区分")
        if kubun:
            kubun_values = str(kubun).split("_")
            if "2" in kubun_values or "4" in kubun_values:
                red_bg = "background-color: #FF0000; color: #FFFFFF"
                styles[idx["部品区分"]] = styles[idx["部品区分"]] + "; " + red_bg if styles[idx["部品区分"]] else red_bg

        return styles

    styled_df = df.style.apply(_highlight_rows, axis=1)
    assembly_details = stats.get("assembly_details", {})
    m_details = stats.get("m_details", {})

    # ---------- 結果テーブル ----------
    st.markdown('<div class="section-header">結果一覧</div>',
                unsafe_allow_html=True)
    st.caption("A番・M番の行を選択すると詳細が表示されます")
    event = st.dataframe(
        styled_df,
        on_select="rerun",
        selection_mode="single-row",
        use_container_width=True,
        height=600,
    )

    # 在庫CD未登録チェック
    missing_zaiko = [r.buhin_bango for r in results if not r.zaiko_cd]
    if missing_zaiko:
        st.error("在庫CDが登録されていません")
        with st.expander(f"対象品番 ({len(missing_zaiko)}件)"):
            st.write(", ".join(missing_zaiko))

    # Excel ダウンロード
    buf = io.BytesIO()
    write_results(results, buf)
    buf.seek(0)

    st.download_button(
        label="Excelダウンロード",
        data=buf,
        file_name="price_results.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
    )

    # ---------- 行選択 → 詳細表示 ----------
    selected_pn = None
    if event.selection.rows:
        row_idx = event.selection.rows[0]
        if row_idx < len(results):
            selected_pn = results[row_idx].buhin_bango

    # ---------- A番 構成部品詳細 ----------
    if selected_pn and _is_a_part(selected_pn) and selected_pn in assembly_details:
        st.markdown("---")
        st.markdown('<div class="section-header">A番 構成部品詳細</div>',
                    unsafe_allow_html=True)
        detail = assembly_details[selected_pn]

        if detail.has_null_data:
            st.markdown(f"### :red[{selected_pn} の構成部品]")
            st.warning("構成部品に標準単価が取得できないものがあります。")
        else:
            st.markdown(f"### {selected_pn} の構成部品")

        # サマリー情報
        col1, col2, col3, col4 = st.columns(4)
        col1.metric("単価合計", f"{detail.buhin_total:,.0f}")
        col2.metric("H仕切合計", f"{detail.h_sikiri_total:,}")
        col3.metric("工数×チャージ", f"{detail.kousuu_x_charge:,.0f}")
        col4.metric("社外組立費", f"{detail.kumitate_gaichuhi:,.0f}")

        col5, col6, col7, col8 = st.columns(4)
        col5.metric("工数(分)", f"{detail.kousuu:,.1f}")
        col6.metric("単価合計+組立費", f"{detail.genka_total:,.0f}")
        col7.metric("組立場所", detail.assembly_place or "-")
        a_result = next((r for r in results if r.buhin_bango == selected_pn), None)
        col8.metric("A番H仕切", f"{a_result.h_sikiri:,}" if a_result and a_result.h_sikiri else "-")

        # 構成部品テーブル（行選択でM番工程詳細を表示）
        comp_rows = []
        for comp in detail.components:
            comp_rows.append({
                "部品番号": comp.buhin_bango,
                "員数": int(comp.inzuu),
                "単価": float(comp.tanka) if comp.tanka is not None else None,
                "H仕切り": comp.h_sikiri,
                "H仕切り×員数": comp.h_sikiri_x_inzuu,
                "部品名": comp.buhin_name,
            })
        if comp_rows:
            comp_df = pd.DataFrame(comp_rows)

            def _highlight_null_comp(row_s: pd.Series) -> list[str]:
                if row_s.get("単価") is None or row_s.get("H仕切り") is None:
                    return ["color: #FF0000"] * len(row_s)
                return [""] * len(row_s)

            styled_comp = comp_df.style.apply(_highlight_null_comp, axis=1)

            # M番子部品がある場合は行選択可能に
            has_m_child = any(_is_m_part(c.buhin_bango) and c.buhin_bango in m_details
                             for c in detail.components)
            if has_m_child:
                st.caption("M番の行を選択すると工程詳細が表示されます")
                comp_event = st.dataframe(
                    styled_comp,
                    on_select="rerun",
                    selection_mode="single-row",
                    use_container_width=True,
                    hide_index=False,
                )
                # M番子部品選択時 → 工程詳細表示
                if comp_event.selection.rows:
                    comp_idx = comp_event.selection.rows[0]
                    if comp_idx < len(detail.components):
                        comp_pn = detail.components[comp_idx].buhin_bango
                        if _is_m_part(comp_pn) and comp_pn in m_details:
                            selected_pn = comp_pn  # M番詳細表示に流す
            else:
                st.dataframe(styled_comp, use_container_width=True, hide_index=False)
        else:
            st.info("構成部品がありません。")

    # ---------- M番 工程内容詳細 ----------
    if selected_pn and _is_m_part(selected_pn):
        st.markdown("---")
        st.markdown('<div class="section-header">M番 工程内容詳細</div>',
                    unsafe_allow_html=True)

        # 結果一覧と同じ hyotanka から外注コストを表示
        pr = next((r for r in results if r.buhin_bango == selected_pn), None)
        if pr:
            col1, col2, col3 = st.columns(3)
            col1.metric("外注コスト", f"{pr.gaikote_cost:,.0f}" if pr.gaikote_cost is not None else "-")
            col2.metric("内作コスト", f"{pr.naikote_cost:,.0f}" if pr.naikote_cost is not None else "-")
            col3.metric("購入コスト", f"{pr.konyu_cost:,.0f}" if pr.konyu_cost is not None else "-")

        if selected_pn in m_details:
            detail_m = m_details[selected_pn]
            st.markdown(f"### {selected_pn} の工程内容")
            if detail_m.buhi_mei:
                st.caption(f"部品名: {detail_m.buhi_mei}")

            proc_rows = []
            for proc in detail_m.processes:
                row_data = {
                    "工程順": proc.kote_jun,
                    "工程": proc.koutei,
                    "課": proc.ka,
                    "班": proc.han,
                    "業者": proc.gyusya,
                }
                # 業者コストはデータがある場合のみ表示
                if proc.gyusyacost is not None:
                    row_data["業者コスト"] = float(proc.gyusyacost)
                row_data.update({
                    "段取時間": float(proc.in_plan_t) if proc.in_plan_t is not None else None,
                    "LOT付帯": float(proc.lot_inc_t) if proc.lot_inc_t is not None else None,
                    "部品付帯": float(proc.buh_inc_t) if proc.buh_inc_t is not None else None,
                    "加工サイクル": float(proc.kakou_cycle_t) if proc.kakou_cycle_t is not None else None,
                    "機人": proc.kijin_flg,
                    "材料費": float(proc.zairyo_cost) if proc.zairyo_cost is not None else None,
                })
                proc_rows.append(row_data)
            if proc_rows:
                proc_df = pd.DataFrame(proc_rows)
                st.dataframe(proc_df, use_container_width=True, hide_index=True)
            else:
                st.info("工程データがありません。")
        else:
            st.info(f"{selected_pn} の工程データがHONPSに登録されていません。")
