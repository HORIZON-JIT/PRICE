"""設定ファイル読み込み."""
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import yaml
from openpyxl import load_workbook


@dataclass
class DbConfig:
    """データベース接続設定."""
    user: str
    password: str
    dsn: str
    pool_min: int = 2
    pool_max: int = 10


@dataclass
class MBandConfig:
    """M番の価格帯別掛率設定."""
    band1_threshold: Decimal
    band2_threshold: Decimal
    band3_threshold: Decimal
    rate_M1: Decimal
    rate_M2: Decimal
    rate_M3: Decimal


@dataclass
class PriceChainConfig:
    """価格チェーン設定."""
    hi_var1: Decimal
    hi_var2: Decimal
    hi_var3: Decimal
    kari_var1: Decimal
    kari_var2: Decimal
    kari_var3: Decimal  # 4番以外
    kari_var4: Decimal  # 4番以外
    dealer_var1: Decimal
    jyoudai_band1: Decimal
    jyoudai_band2: Decimal
    jyoudai_rate1: Decimal
    jyoudai_rate2: Decimal
    jyoudai_rate3: Decimal


@dataclass
class RateConfig:
    """全掛率設定."""
    m_band: MBandConfig
    rate_4: Decimal
    rate_e: Decimal
    rate_a: Decimal
    rate_l: Decimal
    rate_cv: Decimal
    rate_um: Decimal
    rate_p: Decimal
    up_rate: Decimal
    charge_rate: Decimal
    price_chain: PriceChainConfig
    price_comparison_rate: Decimal = Decimal("1.3")  # ECO H仕切り比較用掛率 (VBA: Y1)


@dataclass
class AppConfig:
    """アプリケーション全体設定."""
    eco_db: DbConfig
    honps_db: DbConfig
    rates: RateConfig
    chunk_size: int = 999
    max_depth: int = 10
    history_days: int = 365
    a_assembly_mode: str = "simple"  # "simple"=簡易式, "kousuu"=工数反映式, "kousuu_2026"=工数反映式_2026新型


def _to_decimal(val) -> Decimal:
    if val is None:
        raise ValueError("掛率が未設定(null)です。config/rates.yamlを確認してください。")
    return Decimal(str(val))


def _cell_val(ws, cell_ref) -> Decimal:
    """Excelシートのセル値をDecimalで取得する.

    Args:
        ws: openpyxlのワークシート
        cell_ref: セル参照 (例: "A3", "F10")
    """
    val = ws[cell_ref].value
    if val is None:
        raise ValueError(f"テーブルシートのセル {cell_ref} が空です。")
    return Decimal(str(val))


def load_rates_from_excel(excel_path: str | Path, sheet_name: str = "テーブル") -> RateConfig:
    """Excelの「テーブル」シートから掛率を直接読み込む.

    VBAと同じセル位置から値を取得する:
      A3: UP, B3: チャージ
      D5: M番価格帯3下限, E3: 価格帯1上限, E4: 価格帯2上限
      F3~F5: M番掛率M1~M3
      F10: 4番掛率, F19: E番掛率, F23: P番掛率, F27: A番掛率
      F32: L番掛率, F36: CV番掛率, F40: UM番掛率
      F51~F53: HI仕切り変数, F57~F60: 仮上代変数
      F63: D仕切り変数
      E71~E72: 上代価格帯, F71~F73: 上代掛率
    """
    wb = load_workbook(excel_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"シート '{sheet_name}' が見つかりません。"
            f"利用可能なシート: {wb.sheetnames}"
        )
    ws = wb[sheet_name]

    m_band = MBandConfig(
        band1_threshold=_cell_val(ws, "E3"),
        band2_threshold=_cell_val(ws, "E4"),
        band3_threshold=_cell_val(ws, "D5"),
        rate_M1=_cell_val(ws, "F3"),
        rate_M2=_cell_val(ws, "F4"),
        rate_M3=_cell_val(ws, "F5"),
    )

    price_chain = PriceChainConfig(
        hi_var1=_cell_val(ws, "F51"),
        hi_var2=_cell_val(ws, "F52"),
        hi_var3=_cell_val(ws, "F53"),
        kari_var1=_cell_val(ws, "F57"),
        kari_var2=_cell_val(ws, "F58"),
        kari_var3=_cell_val(ws, "F59"),
        kari_var4=_cell_val(ws, "F60"),
        dealer_var1=_cell_val(ws, "F63"),
        jyoudai_band1=_cell_val(ws, "E71"),
        jyoudai_band2=_cell_val(ws, "E72"),
        jyoudai_rate1=_cell_val(ws, "F71"),
        jyoudai_rate2=_cell_val(ws, "F72"),
        jyoudai_rate3=_cell_val(ws, "F73"),
    )

    # Y1: ECO H仕切り比較用掛率 (セルが空ならデフォルト1.3)
    y1_val = ws["Y1"].value
    pcr = Decimal(str(y1_val)) if y1_val is not None else Decimal("1.3")

    rate_config = RateConfig(
        m_band=m_band,
        rate_4=_cell_val(ws, "F10"),
        rate_e=_cell_val(ws, "F19"),
        rate_a=_cell_val(ws, "F27"),
        rate_l=_cell_val(ws, "F32"),
        rate_cv=_cell_val(ws, "F36"),
        rate_um=_cell_val(ws, "F40"),
        rate_p=_cell_val(ws, "F23"),
        up_rate=_cell_val(ws, "A3"),
        charge_rate=_cell_val(ws, "B3"),
        price_chain=price_chain,
        price_comparison_rate=pcr,
    )

    wb.close()
    return rate_config


def _load_rates_from_yaml(rates_path: str | Path) -> RateConfig:
    """YAMLファイルから掛率を読み込む（従来方式）."""
    with open(rates_path, encoding="utf-8") as f:
        rates = yaml.safe_load(f)

    ts = rates["h_sikiri"]
    m = ts["M"]
    mfg = rates["manufacturing"]
    pc = rates["price_chain"]

    m_band = MBandConfig(
        band1_threshold=_to_decimal(m["band1_threshold"]),
        band2_threshold=_to_decimal(m["band2_threshold"]),
        band3_threshold=_to_decimal(m["band3_threshold"]),
        rate_M1=_to_decimal(m["rate_M1"]),
        rate_M2=_to_decimal(m["rate_M2"]),
        rate_M3=_to_decimal(m["rate_M3"]),
    )

    price_chain = PriceChainConfig(
        hi_var1=_to_decimal(pc["hi_sikiri"]["var1"]),
        hi_var2=_to_decimal(pc["hi_sikiri"]["var2"]),
        hi_var3=_to_decimal(pc["hi_sikiri"]["var3"]),
        kari_var1=_to_decimal(pc["kari_jyoudai"]["var1"]),
        kari_var2=_to_decimal(pc["kari_jyoudai"]["var2"]),
        kari_var3=_to_decimal(pc["kari_jyoudai"]["var3"]),
        kari_var4=_to_decimal(pc["kari_jyoudai"]["var4"]),
        dealer_var1=_to_decimal(pc["dealer_sikiri"]["var1"]),
        jyoudai_band1=_to_decimal(pc["jyoudai"]["band1_threshold"]),
        jyoudai_band2=_to_decimal(pc["jyoudai"]["band2_threshold"]),
        jyoudai_rate1=_to_decimal(pc["jyoudai"]["rate1"]),
        jyoudai_rate2=_to_decimal(pc["jyoudai"]["rate2"]),
        jyoudai_rate3=_to_decimal(pc["jyoudai"]["rate3"]),
    )

    pcr = rates.get("price_comparison_rate", 1.3)

    return RateConfig(
        m_band=m_band,
        rate_4=_to_decimal(ts["4"]),
        rate_e=_to_decimal(ts["E"]),
        rate_a=_to_decimal(ts["A"]),
        rate_l=_to_decimal(ts["L"]),
        rate_cv=_to_decimal(ts["CV"]),
        rate_um=_to_decimal(ts["UM"]),
        rate_p=_to_decimal(ts["P"]),
        up_rate=_to_decimal(mfg["up_rate"]),
        charge_rate=_to_decimal(mfg["charge_rate"]),
        price_chain=price_chain,
        price_comparison_rate=_to_decimal(pcr),
    )


def load_config(
    settings_path: str | Path = "config/settings.yaml",
    rates_path: str | Path | None = "config/rates.yaml",
    rates_excel_path: str | Path | None = None,
    rates_sheet_name: str = "テーブル",
    a_assembly_mode: str = "simple",
) -> AppConfig:
    """設定ファイルを読み込んでAppConfigを返す.

    掛率は以下の優先順位で読み込む:
      1. rates_excel_path が指定されている → Excelの「テーブル」シートから読み込み
      2. それ以外 → rates_path (YAML) から読み込み

    Args:
        settings_path: DB接続等の設定YAMLパス
        rates_path: 掛率設定YAMLパス (Excelが指定されていない場合に使用)
        rates_excel_path: 掛率が入ったExcelファイルパス (VBAの「テーブル」シート)
        rates_sheet_name: Excelのシート名 (default: "テーブル")
        a_assembly_mode: A番組立金額の計算方式 ("simple"=簡易式, "kousuu"=工数反映式)
    """
    with open(settings_path, encoding="utf-8") as f:
        settings = yaml.safe_load(f)

    db = settings["database"]
    eco_db = DbConfig(**db["eco"])
    honps_db = DbConfig(**db["honps"])

    # 掛率の読み込み: Excel優先
    if rates_excel_path is not None:
        rate_config = load_rates_from_excel(rates_excel_path, rates_sheet_name)
    else:
        rate_config = _load_rates_from_yaml(rates_path)

    batch = settings.get("batch", {})
    return AppConfig(
        eco_db=eco_db,
        honps_db=honps_db,
        rates=rate_config,
        chunk_size=batch.get("chunk_size", 999),
        max_depth=batch.get("max_depth", 10),
        history_days=batch.get("manufacturing", {}).get("history_days", 365),
        a_assembly_mode=a_assembly_mode,
    )
