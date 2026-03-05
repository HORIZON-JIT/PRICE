"""Excel出力モジュール.

VBAの取込データ作成()に対応。計算結果をExcelファイルに書き出す。
"""
import io
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, numbers

from price.models.price_result import PriceResult

# セルの色設定
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CYAN_FILL = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
HEADER_FONT = Font(bold=True)

# 出力列定義
COLUMNS = [
    ("品番", "buhin_bango"),
    ("標準単価", "standard_price"),
    ("T仕切り", "t_sikiri"),
    ("掛率", "kakeru"),
    ("HI仕切り", "hi_sikiri"),
    ("仮上代", "kari_jyoudai"),
    ("D仕切り", "dealer_sikiri"),
    ("上代", "jyoudai"),
    ("ECO H仕切り", "h_sikiri_eco"),
    ("ECO H仕切り(調整後)", "h_sikiri_eco_adjusted"),
    ("内作コスト", "naikote_cost"),
    ("外注コスト", "gaikote_cost"),
    ("購入コスト", "konyu_cost"),
    ("第1工程", "first_process"),
    ("部品区分", "buhin_kubun"),
    ("価格比較", "price_comparison_flag"),
    ("NULLデータ", "has_null_data"),
]


def write_results(results: list[PriceResult], output_path: str | Path) -> Path:
    """計算結果をExcelファイルに出力する.

    Args:
        results: 計算結果リスト
        output_path: 出力ファイルパス

    Returns:
        出力したファイルのパス
    """
    wb = Workbook()

    # 一括シート
    ws = wb.active
    ws.title = "一括"

    # ヘッダー
    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT

    # データ行
    for row_idx, result in enumerate(results, 2):
        for col_idx, (_, attr) in enumerate(COLUMNS, 1):
            value = getattr(result, attr, None)
            if attr == "has_null_data":
                value = "有" if value else ""

            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # 数値列のフォーマット
            if attr in ("standard_price", "naikote_cost", "gaikote_cost",
                         "konyu_cost", "h_sikiri_eco", "kakeru"):
                cell.number_format = "#,##0.00"
            elif attr in ("t_sikiri", "hi_sikiri", "kari_jyoudai",
                           "dealer_sikiri", "jyoudai", "h_sikiri_eco_adjusted"):
                cell.number_format = "#,##0"

        # 色付け: 標準単価が0またはNULLの場合は黄色
        if result.standard_price is None or result.standard_price == 0:
            ws.cell(row=row_idx, column=1).fill = YELLOW_FILL
            ws.cell(row=row_idx, column=2).fill = YELLOW_FILL

        # 色付け: T仕切り >= ECO調整後 → 黄色、T仕切り < ECO H仕切り → 水色
        if result.price_comparison_flag == "高":
            ws.cell(row=row_idx, column=3).fill = YELLOW_FILL
        elif result.price_comparison_flag == "安":
            ws.cell(row=row_idx, column=3).fill = CYAN_FILL

    # 列幅の自動調整
    for col_idx, (header, _) in enumerate(COLUMNS, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(
            len(header) * 2 + 2, 12
        )

    # 取込データシート (ECOインポート用)
    _write_import_sheet(wb, results)

    if isinstance(output_path, io.IOBase):
        wb.save(output_path)
        return output_path
    output_path = Path(output_path)
    wb.save(output_path)
    return output_path


def _write_import_sheet(wb: Workbook, results: list[PriceResult]) -> None:
    """ECO取り込み用データシートを作成する."""
    ws = wb.create_sheet("取込データ")
    headers = ["得意先CD", "得意先名称", "在庫CD", "在庫名称", "コンフィグ",
               "品番", "品目名称", "開始日時", "HI仕切り", "インター仕切り",
               "ディーラー仕切り", "上代", "備考"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT

    row = 2
    for r in results:
        if r.t_sikiri is None:
            continue
        # T10000行
        ws.cell(row=row, column=1, value="T10000")
        ws.cell(row=row, column=6, value=r.buhin_bango)
        ws.cell(row=row, column=9, value=r.hi_sikiri)
        ws.cell(row=row, column=10, value=r.kari_jyoudai)
        ws.cell(row=row, column=11, value=r.dealer_sikiri)
        ws.cell(row=row, column=12, value=r.jyoudai)
        row += 1
        # T20000行
        ws.cell(row=row, column=1, value="T20000")
        ws.cell(row=row, column=6, value=r.buhin_bango)
        ws.cell(row=row, column=9, value=r.hi_sikiri)
        ws.cell(row=row, column=10, value=r.kari_jyoudai)
        ws.cell(row=row, column=11, value=0)
        ws.cell(row=row, column=12, value=r.jyoudai)
        row += 1
