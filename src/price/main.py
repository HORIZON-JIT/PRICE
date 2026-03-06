"""価格計算バッチ処理のメインエントリーポイント.

VBAの一括()に対応する2フェーズ処理:
  Phase 1: データ一括取得
  Phase 2: 一括計算 → Excel出力
"""
from __future__ import annotations

import argparse
import sys
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import load_workbook

from price.calc.dispatcher import PriceDispatcher
from price.config import AppConfig, load_config
from price.db.eco_repo import EcoRepo
from price.db.honps_repo import HonpsRepo
from price.db.pool import PoolManager
from price.export.excel_writer import write_results
from price.models.enums import PartPrefix, classify_prefix
from price.models.price_result import PriceResult


def read_input_parts(input_path: str | Path) -> list[str]:
    """Excelファイルから品番リストを読み込む.

    A列に品番が入っている前提（1行目はヘッダー）。
    """
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    parts = []
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        val = row[0]
        if val is not None and str(val).strip():
            parts.append(str(val).strip())
    wb.close()
    return parts


def prefetch_data(
    part_numbers: list[str],
    config: AppConfig,
    on_progress: callable | None = None,
) -> dict:
    """Phase 1: 全データを一括プリフェッチする.

    VBAでは各行で個別にDB接続・クエリしていた処理を、
    プレフィックス別にグルーピングして一括取得する。
    独立したDBクエリは ThreadPoolExecutor で並列実行する。
    """
    # 進捗管理: Wave1 + Wave2 + 計算完了 = 合計3ステップ
    total_steps = 3
    current_step = 0

    def _step(msg: str) -> None:
        nonlocal current_step
        current_step += 1
        print(f"  {msg}")
        if on_progress:
            on_progress(current_step, total_steps, msg)

    data: dict = {}

    # プレフィックス別に分類
    groups: dict[PartPrefix, list[str]] = defaultdict(list)
    for pn in part_numbers:
        try:
            prefix = classify_prefix(pn)
            groups[prefix].append(pn)
        except ValueError:
            pass

    print(f"  部品数: {len(part_numbers)}")
    for prefix, pns in groups.items():
        print(f"    {prefix.value}番: {len(pns)}件")

    non_a_parts = [pn for pn in part_numbers
                   if classify_prefix(pn) != PartPrefix.A]
    a_parts = groups.get(PartPrefix.A, [])
    four_parts = groups.get(PartPrefix.FOUR, [])
    um_parts = groups.get(PartPrefix.UM, [])

    # ── Wave 1: 独立クエリを並列実行 ──
    _step("データ一括取得中 (Wave 1)...")
    with ThreadPoolExecutor(max_workers=5) as executor:
        fut_hyotanka = executor.submit(HonpsRepo.fetch_hyotanka, non_a_parts)
        fut_a_comp = (
            executor.submit(EcoRepo.fetch_a_components, a_parts)
            if a_parts else None
        )
        fut_kakaku = (
            executor.submit(EcoRepo.fetch_kakakuhyou, four_parts)
            if four_parts else None
        )
        fut_shohin = executor.submit(EcoRepo.fetch_shohin_buhin, part_numbers)
        fut_kubun = executor.submit(EcoRepo.fetch_buhin_kubun, part_numbers)

        # 結果回収
        data["hyotanka"] = fut_hyotanka.result()
        data["_hyotanka_debug"] = data["hyotanka"].pop("__debug__", [])

        if fut_a_comp is not None:
            data["a_components"] = fut_a_comp.result()

        if fut_kakaku is not None:
            data["kakakuhyou"] = fut_kakaku.result()
            data.setdefault("_kakaku_debug", [])
            data["_kakaku_debug"] = data["kakakuhyou"].pop("__debug__", [])

        data["shohin_buhin"] = fut_shohin.result()
        data["buhin_kubun"] = fut_kubun.result()

    # ── Wave 2: Wave1 の結果に依存するクエリを並列実行 ──
    _step("データ一括取得中 (Wave 2)...")
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures: list[tuple[str, object]] = []

        # A番構成部品の標準単価 + 組立費 + 工数
        if a_parts and "a_components" in data:
            component_pns = set()
            for comps in data["a_components"].values():
                for c in comps:
                    component_pns.add(c.buhin_bango)
            if component_pns:
                futures.append(("comp_hyotanka",
                    executor.submit(HonpsRepo.fetch_hyotanka,
                                    list(component_pns))))
            futures.append(("a_assembly_cost",
                executor.submit(EcoRepo.fetch_a_assembly_cost, a_parts)))
            futures.append(("a_assembly_kousuu",
                executor.submit(EcoRepo.fetch_a_assembly_kousuu, a_parts)))

        # M番工程詳細 (トップレベル + A番構成部品内のM番)
        m_parts = set(groups.get(PartPrefix.M, []))
        if a_parts and "a_components" in data:
            for comps in data["a_components"].values():
                for c in comps:
                    try:
                        if classify_prefix(c.buhin_bango) == PartPrefix.M:
                            m_parts.add(c.buhin_bango)
                    except ValueError:
                        pass
        if m_parts:
            futures.append(("m_buhin",
                executor.submit(HonpsRepo.fetch_m_buhin, list(m_parts))))

        # 外貨レート（バッチ取得）
        if four_parts and "kakakuhyou" in data:
            currencies = set()
            for kk in data["kakakuhyou"].values():
                if kk.tori_tuuka_tani_kbn and kk.tori_tuuka_tani_kbn != "JPY":
                    currencies.add(kk.tori_tuuka_tani_kbn)
            if currencies:
                futures.append(("fx_rates",
                    executor.submit(EcoRepo.fetch_rates, list(currencies))))

        # UM番 H仕切り
        if um_parts:
            futures.append(("um_shohin",
                executor.submit(EcoRepo.fetch_um_h_sikiri, um_parts)))

        # 結果回収
        for key, fut in futures:
            result = fut.result()
            if key == "comp_hyotanka":
                data["hyotanka"].update(result)
            elif key == "um_shohin":
                for pn, sb in result.items():
                    existing = data["shohin_buhin"].get(pn)
                    if existing is None or existing.h_sikiri is None:
                        data["shohin_buhin"][pn] = sb
            elif key == "fx_rates":
                data["fx_rates"] = result
            else:
                data[key] = result

    return data


def process_parts(
    part_numbers: list[str],
    config: AppConfig,
    on_progress: callable | None = None,
) -> tuple[list[PriceResult], dict]:
    """コア処理パイプライン（UI・CLI両対応）.

    Args:
        part_numbers: 品番リスト
        config: AppConfig（掛率・DB設定込み）
        on_progress: 進捗コールバック(step, total, message)

    Returns:
        (results, stats) stats には fetch_time, calc_time, total, null_count
    """
    start = time.time()
    data = prefetch_data(part_numbers, config, on_progress=on_progress)
    fetch_time = time.time() - start

    # 計算ステップ (step 3/3)
    if on_progress:
        on_progress(3, 3, "計算実行中...")
    start = time.time()
    dispatcher = PriceDispatcher(config.rates)
    results = dispatcher.calculate_batch(part_numbers, data)
    calc_time = time.time() - start

    null_count = sum(1 for r in results if r.has_null_data)
    # M番詳細: dispatcher経由(トップレベル) + data直接(A番子部品)をマージ
    m_details = dict(dispatcher.get_m_details())
    m_details.update(data.get("m_buhin", {}))
    stats = {
        "fetch_time": fetch_time,
        "calc_time": calc_time,
        "total": len(results),
        "null_count": null_count,
        "assembly_details": dispatcher.get_assembly_details(),
        "m_details": m_details,
        "hyotanka_debug": data.get("_hyotanka_debug", []),
        "kakaku_debug": data.get("_kakaku_debug", []),
    }
    return results, stats


def run_batch(input_path: str, output_path: str,
              settings_path: str = "config/settings.yaml",
              rates_path: str = "config/rates.yaml",
              rates_excel_path: str | None = None,
              rates_sheet_name: str = "テーブル") -> None:
    """バッチ処理のメインフロー."""
    print("=" * 60)
    print("価格計算バッチ処理")
    print("=" * 60)

    # 設定読み込み
    print("\n[1/5] 設定ファイル読み込み...")
    if rates_excel_path:
        print(f"  掛率: {rates_excel_path} (シート: {rates_sheet_name})")
    else:
        print(f"  掛率: {rates_path}")
    config = load_config(settings_path, rates_path,
                         rates_excel_path=rates_excel_path,
                         rates_sheet_name=rates_sheet_name)

    # DB接続プール初期化
    print("[2/5] データベース接続...")
    PoolManager.init(config.eco_db, config.honps_db)

    try:
        # 入力読み込み
        print(f"[3/5] 入力ファイル読み込み: {input_path}")
        part_numbers = read_input_parts(input_path)
        if not part_numbers:
            print("エラー: 品番が見つかりません。")
            return

        # Phase 1 + 2: データ取得 → 計算
        print(f"[4/5] データ取得中...")
        results, stats = process_parts(part_numbers, config)
        print(f"  データ取得完了: {stats['fetch_time']:.1f}秒")
        print(f"[5/5] 計算完了: {stats['calc_time']:.1f}秒")

        # Excel出力
        out = write_results(results, output_path)
        print(f"\n結果出力: {out}")
        print(f"  処理件数: {stats['total']}")
        if stats["null_count"]:
            print(f"  NULLデータあり: {stats['null_count']}件")
        print(f"  合計処理時間: {stats['fetch_time'] + stats['calc_time']:.1f}秒")

    finally:
        PoolManager.close()

    print("\n完了")


def main():
    parser = argparse.ArgumentParser(
        description="価格計算バッチ処理システム",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
使用例:
  price --input parts.xlsx --output result.xlsx
  price -i parts.xlsx -o result.xlsx --settings config/settings.yaml --rates config/rates.yaml
        """,
    )
    parser.add_argument("-i", "--input", required=True,
                        help="入力Excelファイル (A列に品番)")
    parser.add_argument("-o", "--output", required=True,
                        help="出力Excelファイル")
    parser.add_argument("--settings", default="config/settings.yaml",
                        help="DB接続設定ファイル (default: config/settings.yaml)")
    parser.add_argument("--rates", default="config/rates.yaml",
                        help="掛率設定YAMLファイル (default: config/rates.yaml)")
    parser.add_argument("--rates-excel",
                        help="掛率が入ったExcelファイル (「テーブル」シートから読み込み)")
    parser.add_argument("--rates-sheet", default="テーブル",
                        help="Excelの掛率シート名 (default: テーブル)")

    args = parser.parse_args()
    run_batch(args.input, args.output, args.settings, args.rates,
              rates_excel_path=args.rates_excel,
              rates_sheet_name=args.rates_sheet)


if __name__ == "__main__":
    main()
